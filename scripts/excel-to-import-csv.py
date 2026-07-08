#!/usr/bin/env python3
"""Convert the NPPA LPD 2026 Sponsorship Master workbook into a CSV that the CRM's
in-app importer (/admin/events/[id]/import) can ingest.

Usage:
    python3 scripts/excel-to-import-csv.py <input.xlsx> <output.csv>

Only the "Master List" sheet is read. Status/priority are lowercased to the CRM's
enum vocabulary; tiers pass through (matched case-insensitively by the importer);
the single "Best POC" / "Alt Contact Name" full-name fields are split into
first/last for contact 1 (primary) and contact 2. Output column names match the
importer's recognized headers. The output CSV contains contact PII — do not commit it.
"""
import csv
import sys
from datetime import date, datetime

import openpyxl

SHEET_NAME = "Master List"

STATUS_MAP = {
    "prospect": "prospect",
    "not contacted": "prospect",
    "contacted": "contacted",
    "engaged": "engaged",
    "proposal sent": "proposal_sent",
    "negotiating": "negotiating",
    "committed": "committed",
    "confirmed": "confirmed",
    "declined": "declined",
    "past sponsor": "past_sponsor",
}

PRIORITY_MAP = {"high": "high", "medium": "medium", "low": "low"}

# Map the workbook's free-text owner names to a value the CRM importer resolves
# (it matches users by full name, first name, or email local-part).
OWNER_ALIASES = {"michael": "Mike Thorn", "mike": "Mike Thorn"}

# Sponsorship tiers that map to a real tier; anything else (e.g. "Unknown") -> blank.
VALID_TIERS = {"platinum", "gold", "silver", "bronze"}

# Maps each importer output column to the workbook column it is sourced from.
# Columns needing transformation (name splits, remaps) are handled in main().
DIRECT_COLUMNS = {
    "website": "Website",
    "industry": "Category",
    "subcategory": "Subcategory",
    "hq_location": "HQ Location",
    "why_they_should_attend": "What They Do / Why They Fit",
    "key_talking_points": "Key Talking Points",
    "email_angle": "Email Angle",
    "sponsorship_hook": "Sponsorship Hook",
    "relationship_notes": "Notes",
    "first_contacted_at": "First Contacted",
    "last_contacted_at": "Last Contacted",
    "contact1_email": "POC Email",
    "contact1_title": "POC Title",
    "contact2_email": "Alt Contact Email",
    "contact2_title": "Alt Contact Title",
}

OUTPUT_COLUMNS = [
    "name",
    "website",
    "industry",
    "subcategory",
    "status",
    "priority",
    "owner",
    "target_tier",
    "hq_location",
    "why_they_should_attend",
    "key_talking_points",
    "email_angle",
    "sponsorship_hook",
    "relationship_notes",
    "first_contacted_at",
    "last_contacted_at",
    "contact1_first_name",
    "contact1_last_name",
    "contact1_email",
    "contact1_title",
    "contact1_phone",
    "contact1_linkedin",
    "contact2_first_name",
    "contact2_last_name",
    "contact2_email",
    "contact2_title",
    "contact2_phone",
    "contact2_linkedin",
]


def s(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    return str(v).strip()


def split_name(full):
    """Split a single full-name string into (first, last)."""
    parts = s(full).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    in_path, out_path = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(in_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    header = [s(h) for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    missing = [c for c in DIRECT_COLUMNS.values() if c not in idx]
    for extra in ("Company", "Status", "Priority", "Target Tier", "Owner",
                  "Best POC", "Alt Contact Name"):
        if extra not in idx:
            missing.append(extra)
    if missing:
        print(f"WARNING: workbook is missing expected columns: {missing}")

    def cell(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out_rows = []
    skipped = 0
    for row in rows[1:]:
        if not any(c not in (None, "") for c in row):
            continue
        name = s(cell(row, "Company"))
        if not name:
            skipped += 1
            continue

        out = {c: "" for c in OUTPUT_COLUMNS}
        out["name"] = name

        for out_col, src_col in DIRECT_COLUMNS.items():
            out[out_col] = s(cell(row, src_col))

        out["status"] = STATUS_MAP.get(s(cell(row, "Status")).lower(), "")
        out["priority"] = PRIORITY_MAP.get(s(cell(row, "Priority")).lower(), "")

        tier_raw = s(cell(row, "Target Tier")).lower()
        out["target_tier"] = tier_raw.capitalize() if tier_raw in VALID_TIERS else ""

        owner_raw = s(cell(row, "Owner"))
        out["owner"] = OWNER_ALIASES.get(owner_raw.lower(), owner_raw)

        out["contact1_first_name"], out["contact1_last_name"] = split_name(
            cell(row, "Best POC")
        )
        out["contact2_first_name"], out["contact2_last_name"] = split_name(
            cell(row, "Alt Contact Name")
        )

        out_rows.append(out)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        w.writeheader()
        w.writerows(out_rows)

    print(f"Wrote {len(out_rows)} rows to {out_path} ({skipped} blank rows skipped).")


if __name__ == "__main__":
    main()
